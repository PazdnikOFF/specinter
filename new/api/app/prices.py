"""Конвейер прайсов: парсинг XLSX по профилю поставщика → supplier_prices (с историей)
→ матчинг по артикулу (exact → analog → fuzzy) → пересчёт offers.

Замена legacy `xml.php`: матчинг по АРТИКУЛУ производителя (а не по коду 1С),
per-supplier профили колонок и наценки, история цен без TRUNCATE."""
import io
import re
from datetime import datetime, timezone
from openpyxl import load_workbook
from psycopg.types.json import Json
from . import db

FUZZY_THRESHOLD = 0.62  # порог pg_trgm для авто-сопоставления-кандидата (в очередь модерации)

# ---------------------------------------------------------------------------
# Автоопределение структуры прайса (у каждого поставщика своя раскладка колонок).
# Ключевые слова заголовков → поля. Отлажено на 10 реальных прайсах поставщиков.
HEADER_KEYS = {
    "article": ["артикул", "articul"],
    "name":    ["наименование", "номенклатура", "название", "описание", "запчаст"],
    "price":   ["цена", "стоимость", "безнал", "суперопт", "прайс"],
    "qty":     ["наличие", "остаток", "количество", "кол-во"],
    "maker":   ["производитель", "бренд", "марка", "изготовитель", "brand"],
    "cross":   ["кросс", "cross"],
    "code":    ["код"],
}
STRONG_PRICE = ["цена", "стоимость", "безнал"]
STRONG_QTY = ["остаток", "наличие", "количество"]
STRONG_NAME = ["наименование", "номенклатура"]
# Доп. колонки для расчёта доставки (вес/габариты/объём). Не влияют на поиск шапки.
EXTRA_KEYS = {
    "weight": ["вес", "масса"],
    "volume": ["объ"],                       # объём / объем
    "dim_l":  ["длина", "длинна", "длину"],
    "dim_w":  ["ширина", "ширину"],
    "dim_h":  ["высота", "высоту"],
}

# Идемпотентная миграция: происхождение offer, уникальность (товар, поставщик),
# кросс-номера в прайсе и срок поставки поставщика.
MIGRATION = """
ALTER TABLE offers ADD COLUMN IF NOT EXISTS source text DEFAULT 'legacy';
CREATE UNIQUE INDEX IF NOT EXISTS ux_offers_prod_sup
  ON offers(product_id, supplier_id) WHERE source='price';
ALTER TABLE supplier_prices ADD COLUMN IF NOT EXISTS cross_numbers text;
ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS delivery_days int;
ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS delivery_note text;
ALTER TABLE supplier_price_profiles ADD COLUMN IF NOT EXISTS auto boolean DEFAULT true;
ALTER TABLE supplier_price_profiles ADD COLUMN IF NOT EXISTS col_cross int;
ALTER TABLE products ADD COLUMN IF NOT EXISTS weight_kg numeric;   -- для расчёта доставки
ALTER TABLE products ADD COLUMN IF NOT EXISTS volume_m3 numeric;
ALTER TABLE orders ADD COLUMN IF NOT EXISTS delivery jsonb;        -- выбранная доставка (ДЛ и т.п.)
ALTER TABLE supplier_prices ADD COLUMN IF NOT EXISTS weight_kg numeric;
ALTER TABLE supplier_prices ADD COLUMN IF NOT EXISTS volume_m3 numeric;
ALTER TABLE supplier_prices ADD COLUMN IF NOT EXISTS raw jsonb;  -- ВСЕ колонки прайса
"""


async def ensure_migrated():
    async with db.pool.connection() as conn:
        await conn.execute(MIGRATION)


def _hnorm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _read_rows(data: bytes, filename: str = "") -> list[list]:
    """Читает первую вкладку в список строк независимо от .xls (BIFF) / .xlsx (zip)."""
    is_xlsx = data[:2] == b"PK" or (filename or "").lower().endswith(".xlsx")
    if is_xlsx:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    import xlrd  # старый формат .xls (OLE)
    wb = xlrd.open_workbook(file_contents=data)
    sh = wb.sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def _merged_header(rows, i, span):
    ncol = max((len(rows[j]) for j in range(i, min(i + span, len(rows)))), default=0)
    out = []
    for c in range(ncol):
        parts = [_hnorm(rows[j][c]) for j in range(i, min(i + span, len(rows)))
                 if c < len(rows[j]) and _hnorm(rows[j][c])]
        out.append(" ".join(parts))
    return out


def detect_profile(rows: list[list], scan: int = 20) -> dict | None:
    """Ищет строку-шапку (возможно многострочную) и маппинг колонок. Нужны article+price."""
    best = None
    for i in range(min(scan, len(rows))):
        for span in (1, 2, 3, 4):
            if i + span > len(rows):
                break
            merged = _merged_header(rows, i, span)
            fields = {f for t in merged for f, keys in HEADER_KEYS.items()
                      if t and any(k in t for k in keys)}
            if "article" in fields and "price" in fields:
                key = (len(fields), -span)
                if best is None or key > best[0]:
                    best = (key, i, span, merged)
    if best is None:
        return None
    _, i, span, merged = best
    header_row = i + span - 1
    col_map, price_c, qty_c, name_c = {}, [], [], []
    for idx, t in enumerate(merged):
        if not t:
            continue
        if t == "артикул":
            matched = "article"
        else:
            matched = next((f for f in ["article", "cross", "maker", "qty", "price", "name", "code"]
                            if any(k in t for k in HEADER_KEYS[f])), None)
        if any(k in t for k in STRONG_QTY):
            matched = "qty"
        elif any(k in t for k in STRONG_PRICE) and matched != "article":
            matched = "price"
        if matched == "price":
            price_c.append((idx, t))
        elif matched == "qty":
            qty_c.append((idx, t))
        elif matched == "name":
            name_c.append((idx, t))
        elif matched and matched not in col_map:
            col_map[matched] = idx

    def num_share(idx):
        n = ok = 0
        for r in rows[header_row + 1: header_row + 300]:
            if idx < len(r) and r[idx] not in (None, ""):
                n += 1
                if _f(r[idx]) is not None:
                    ok += 1
        return ok / n if n else 0

    if price_c:
        strong = [i for i, t in price_c if any(k in t for k in STRONG_PRICE)]
        col_map["price"] = max(strong or [i for i, _ in price_c], key=num_share)
    if qty_c:
        strong = [i for i, t in qty_c if any(k in t for k in STRONG_QTY)]
        col_map["qty"] = strong[0] if strong else qty_c[0][0]
    if name_c:
        strong = [i for i, t in name_c if any(k in t for k in STRONG_NAME)]
        col_map["name"] = strong[0] if strong else name_c[0][0]
    # доп. колонки: вес/габариты/объём (не конфликтуют с основными полями)
    for idx, t in enumerate(merged):
        if not t:
            continue
        for f, keys in EXTRA_KEYS.items():
            if f not in col_map and any(k in t for k in keys):
                col_map[f] = idx
    # оригинальные подписи колонок (для сохранения ВСЕХ данных прайса в raw)
    ncol = max((len(rows[j]) for j in range(i, i + span)), default=0)
    labels = []
    for c in range(ncol):
        parts = [str(rows[j][c]).strip() for j in range(i, i + span)
                 if c < len(rows[j]) and rows[j][c] not in (None, "")]
        labels.append(" ".join(parts))
    return {"header_row": header_row, "cols": col_map, "header_labels": labels}


def parse(data: bytes, filename: str = "", profile: dict | None = None) -> list[dict]:
    """Парсит прайс в строки {article,name,price,qty,maker,external_code,cross}.
    По умолчанию автоопределяет колонки; если profile.auto=False — берёт явные
    col_* / header_rows из профиля поставщика."""
    rows = _read_rows(data, filename)
    prof = profile or {}
    if prof.get("auto", True) or not prof.get("col_article"):
        det = detect_profile(rows)
        if not det:
            raise ValueError("не удалось определить структуру прайса (нет колонок артикул+цена)")
        cols, hr = det["cols"], det["header_row"]
        labels = det.get("header_labels") or []
    else:
        cols = {k: prof.get("col_" + k) for k in
                ("article", "name", "price", "qty", "maker", "code", "cross")}
        cols = {("external_code" if k == "code" else k): v for k, v in cols.items() if v is not None}
        hr = (prof.get("header_rows", 1) or 1) - 1
        labels = [str(v).strip() if v not in (None, "") else "" for v in (rows[hr] if hr < len(rows) else [])]

    def raw_of(row):
        # ВСЕ данные прайса построчно: подпись колонки (или colN) → значение.
        out = {}
        for c, v in enumerate(row):
            if v in (None, ""):
                continue
            key = (labels[c] if c < len(labels) and labels[c] else f"col{c}")
            out[key] = v if isinstance(v, (int, float, str)) else str(v)
        return out

    def cell(row, field):
        idx = cols.get("external_code" if field == "code" else field)
        if idx is None or idx < 0 or idx >= len(row):
            return None
        return row[idx]

    def cnum(row, field):
        idx = cols.get(field)
        if idx is None or idx < 0 or idx >= len(row):
            return None
        return _f(row[idx])

    out = []
    for r in rows[hr + 1:]:
        art = cell(r, "article")
        price = cell(r, "price")
        if _s(art) is None:          # без артикула — строки-разделы/пустые
            continue
        # объём: из колонки «объём» или из габаритов Д×Ш×В (см → м³)
        vol = cnum(r, "volume")
        if vol is None:
            l, w, h = cnum(r, "dim_l"), cnum(r, "dim_w"), cnum(r, "dim_h")
            if l and w and h:
                vol = round(l * w * h / 1_000_000, 4)
        out.append({
            "maker": _s(cell(r, "maker")),
            "name": _s(cell(r, "name")),
            "article": _s(art),
            "price": _f(price),
            "qty": _qty(cell(r, "qty")),
            "external_code": _s(cell(r, "code")),
            "cross": _s(cell(r, "cross")),
            "weight_kg": cnum(r, "weight"),
            "volume_m3": vol,
            "raw": raw_of(r),      # ВСЕ колонки прайса (остатки по городам, внутр. коды и т.д.)
        })
    return out


def detect_columns(data: bytes, filename: str = "") -> dict | None:
    """Для админки: показать, какие колонки распознаны (превью профиля)."""
    return detect_profile(_read_rows(data, filename))


# Обратная совместимость со старым именем.
def parse_xlsx(data: bytes, profile: dict) -> list[dict]:
    return parse(data, "", profile)


async def ingest(supplier_id: int, rows: list[dict]) -> dict:
    """Пишет строки прайса (история), матчит по артикулу, пересчитывает offers."""
    await ensure_migrated()
    batch = datetime.now(timezone.utc)
    async with db.pool.connection() as conn:
        async with conn.cursor() as cur:
            for r in rows:
                await cur.execute(
                    """INSERT INTO supplier_prices
                       (supplier_id, article, external_code, name, maker, price, qty,
                        cross_numbers, weight_kg, volume_m3, raw, received_at)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (supplier_id, r["article"], r["external_code"], r["name"],
                     r["maker"], r["price"], r["qty"], r.get("cross"),
                     r.get("weight_kg"), r.get("volume_m3"),
                     Json(r["raw"]) if r.get("raw") else None, batch))

            # --- матчинг только что загруженной партии ---
            # 1) точное совпадение по нормализованному артикулу
            await cur.execute("""
                UPDATE supplier_prices sp SET matched_product_id=p.id,
                       match_method='exact_article', match_confidence=1.0
                FROM products p
                WHERE sp.supplier_id=%s AND sp.received_at=%s AND sp.matched_product_id IS NULL
                  AND sp.normalized_article IS NOT NULL
                  AND p.normalized_article=sp.normalized_article""", (supplier_id, batch))
            # 2) через таблицу аналогов
            await cur.execute("""
                UPDATE supplier_prices sp SET matched_product_id=a.product_id,
                       match_method='analog', match_confidence=0.8
                FROM analogs a
                WHERE sp.supplier_id=%s AND sp.received_at=%s AND sp.matched_product_id IS NULL
                  AND sp.normalized_article IS NOT NULL
                  AND a.normalized_article=sp.normalized_article""", (supplier_id, batch))
            # 2b) по кросс-номерам из прайса (токены → norm_article → наш артикул)
            await cur.execute(r"""
                UPDATE supplier_prices sp SET matched_product_id=p.id,
                       match_method='cross', match_confidence=0.75
                FROM products p
                WHERE sp.supplier_id=%s AND sp.received_at=%s AND sp.matched_product_id IS NULL
                  AND sp.cross_numbers IS NOT NULL
                  AND p.normalized_article = ANY (
                      SELECT norm_article(t)
                      FROM regexp_split_to_table(sp.cross_numbers, '[;,/\s#]+') AS t
                      WHERE norm_article(t) IS NOT NULL
                  )""", (supplier_id, batch))
            # 3) нечёткий (pg_trgm) — только кандидат в очередь модерации (не в offers)
            await cur.execute("""
                UPDATE supplier_prices sp SET matched_product_id=m.pid,
                       match_method='fuzzy', match_confidence=m.sim
                FROM (
                    SELECT spid, pid, sim FROM (
                        SELECT sp2.id AS spid, p.id AS pid,
                               similarity(p.normalized_article, sp2.normalized_article) AS sim,
                               row_number() OVER (PARTITION BY sp2.id
                                   ORDER BY similarity(p.normalized_article, sp2.normalized_article) DESC) AS rn
                        FROM supplier_prices sp2
                        JOIN products p ON p.normalized_article %% sp2.normalized_article
                        WHERE sp2.supplier_id=%s AND sp2.received_at=%s
                          AND sp2.matched_product_id IS NULL AND sp2.normalized_article IS NOT NULL
                    ) ranked WHERE rn=1
                ) m
                WHERE sp.id=m.spid AND m.sim >= %s""",
                (supplier_id, batch, FUZZY_THRESHOLD))

            # --- пересчёт offers из точных/аналоговых матчей (fuzzy — на модерацию) ---
            await cur.execute("DELETE FROM offers WHERE supplier_id=%s AND source='price'", (supplier_id,))
            await cur.execute("""
                INSERT INTO offers (product_id, supplier_id, article, external_code,
                                    price, qty, in_stock, source, updated_at)
                SELECT DISTINCT ON (sp.matched_product_id)
                       sp.matched_product_id, sp.supplier_id, sp.article, sp.external_code,
                       ceil(sp.price*(1+s.markup_percent/100)/s.price_round)*s.price_round,
                       sp.qty, COALESCE(sp.qty,0) > 0, 'price', now()
                FROM supplier_prices sp JOIN suppliers s ON s.id=sp.supplier_id
                WHERE sp.supplier_id=%s AND sp.received_at=%s
                  AND sp.matched_product_id IS NOT NULL
                  AND sp.match_method IN ('exact_article','analog','cross','manual')
                  AND sp.price IS NOT NULL
                ORDER BY sp.matched_product_id, sp.price ASC
                ON CONFLICT (product_id, supplier_id) WHERE source='price'
                DO UPDATE SET price=EXCLUDED.price, qty=EXCLUDED.qty,
                              in_stock=EXCLUDED.in_stock, article=EXCLUDED.article,
                              external_code=EXCLUDED.external_code, updated_at=now()
            """, (supplier_id, batch))

            stats = {}
            await cur.execute(
                """SELECT COALESCE(match_method,'none') AS mm, count(*) AS n
                   FROM supplier_prices WHERE supplier_id=%s AND received_at=%s GROUP BY 1""",
                (supplier_id, batch))
            for row in await cur.fetchall():
                stats[row["mm"]] = row["n"]
    return {"rows": len(rows), "match": stats}


async def backfill_brands() -> int:
    """Заполняет products.brand (производитель) из самого частого maker в прайсах,
    где сопоставление сработало. Не перезатирает уже заданный brand."""
    async with db.pool.connection() as conn:
        async with conn.cursor() as cur:
            await cur.execute("""
                UPDATE products p SET brand = m.maker
                FROM (
                    SELECT matched_product_id, maker,
                           row_number() OVER (PARTITION BY matched_product_id
                                              ORDER BY count(*) DESC) rn
                    FROM supplier_prices
                    WHERE matched_product_id IS NOT NULL AND maker IS NOT NULL AND maker<>''
                    GROUP BY matched_product_id, maker
                ) m
                WHERE p.id = m.matched_product_id AND m.rn = 1
                  AND (p.brand IS NULL OR p.brand = '')
                RETURNING p.id""")
            rows = await cur.fetchall()
    return len(rows)


async def backfill_dimensions() -> int:
    """Заполняет products.weight_kg / volume_m3 из прайсов (где заданы), не перезатирая."""
    async with db.pool.connection() as conn:
        async with conn.cursor() as cur:
            await cur.execute("""
                UPDATE products p SET
                    weight_kg = COALESCE(p.weight_kg, m.weight_kg),
                    volume_m3 = COALESCE(p.volume_m3, m.volume_m3)
                FROM (
                    SELECT matched_product_id,
                           MAX(weight_kg) AS weight_kg, MAX(volume_m3) AS volume_m3
                    FROM supplier_prices
                    WHERE matched_product_id IS NOT NULL
                      AND (weight_kg IS NOT NULL OR volume_m3 IS NOT NULL)
                    GROUP BY matched_product_id
                ) m
                WHERE p.id = m.matched_product_id
                  AND (p.weight_kg IS NULL OR p.volume_m3 IS NULL)
                RETURNING p.id""")
            rows = await cur.fetchall()
    return len(rows)


async def recompute_offers(supplier_id: int) -> int:
    """Пересчитать offers поставщика из ПОСЛЕДНЕЙ партии прайса с ТЕКУЩЕЙ наценкой.
    Вызывается при изменении markup_percent / price_round в карточке поставщика."""
    async with db.pool.connection() as conn:
        async with conn.cursor() as cur:
            await cur.execute("DELETE FROM offers WHERE supplier_id=%s AND source='price'", (supplier_id,))
            await cur.execute("""
                INSERT INTO offers (product_id, supplier_id, article, external_code,
                                    price, qty, in_stock, source, updated_at)
                SELECT DISTINCT ON (sp.matched_product_id)
                       sp.matched_product_id, sp.supplier_id, sp.article, sp.external_code,
                       ceil(sp.price*(1+s.markup_percent/100)/s.price_round)*s.price_round,
                       sp.qty, COALESCE(sp.qty,0) > 0, 'price', now()
                FROM supplier_prices sp JOIN suppliers s ON s.id=sp.supplier_id
                WHERE sp.supplier_id=%s
                  AND sp.received_at = (SELECT max(received_at) FROM supplier_prices WHERE supplier_id=%s)
                  AND sp.matched_product_id IS NOT NULL
                  AND sp.match_method IN ('exact_article','analog','cross','manual')
                  AND sp.price IS NOT NULL
                ORDER BY sp.matched_product_id, sp.price ASC
            """, (supplier_id, supplier_id))
            await cur.execute("SELECT count(*) AS n FROM offers WHERE supplier_id=%s AND source='price'",
                              (supplier_id,))
            row = await cur.fetchone()
    return row["n"]


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _f(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return None


def _qty(v):
    """Гибкое наличие: число → как есть; '<10'/'>10'/'есть'/'в наличии' → в наличии."""
    if v in (None, ""):
        return None
    f = _f(v)
    if f is not None:
        return f
    s = str(v).strip().lower()
    if not s:
        return None
    m = re.search(r"\d+", s)
    if any(sym in s for sym in ("<", ">", "есть", "налич", "много", "+", "да")):
        return float(m.group()) if m else 1.0
    return float(m.group()) if m else None
